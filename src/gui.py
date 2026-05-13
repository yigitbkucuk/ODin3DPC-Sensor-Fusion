import sys, os, cv2, traceback
import numpy as np
import open3d as o3d
from PyQt6.QtWidgets import *
from PyQt6.QtCore import *
from PyQt6.QtGui import *
import pyqtgraph.opengl as gl
from src.processor import process_frame, detect_objects_2d

class DataLoaderThread(QThread):
    finished_signal = pyqtSignal(object, object, object, str)
    error_signal = pyqtSignal(str)

    def __init__(self, path):
        super().__init__()
        self.path = path

    def run(self):
        try:
            pcd, bboxes, meta = process_frame(self.path)

            if pcd is None or len(np.asarray(pcd.points)) == 0:
                self.error_signal.emit("LiDAR verisi islenemedi veya dosya bos!")
                return

            points_np = np.ascontiguousarray(np.asarray(pcd.points, dtype=np.float32).copy())

            safe_bboxes = []
            for b in bboxes:
                safe_bboxes.append({
                    'min_bound': np.ascontiguousarray(np.asarray(b.get_min_bound(), dtype=np.float32).copy()),
                    'max_bound': np.ascontiguousarray(np.asarray(b.get_max_bound(), dtype=np.float32).copy()),
                    'center': np.ascontiguousarray(np.asarray(b.get_center(), dtype=np.float32).copy()),
                    'extent': np.ascontiguousarray(np.asarray(b.get_extent(), dtype=np.float32).copy()),
                    'color': getattr(b, 'color', [1.0, 1.0, 1.0])
                })

            img_path = self.path.replace("velodyne", "image_2").replace(".bin", ".png")

            self.finished_signal.emit(points_np, safe_bboxes, meta, img_path)
        except Exception as e:
            self.error_signal.emit(f"HATA OLUŞTU: {str(e)}")


class CircularProgress(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.value = 0
        self.setFixedSize(200, 200)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        pen = QPen()
        pen.setWidth(10)
        pen.setColor(QColor("#161b22"))
        painter.setPen(pen)
        painter.drawEllipse(10, 10, 180, 180)
        pen.setColor(QColor("#58a6ff"))
        pen.setCapStyle(Qt.PenCapStyle.RoundCap)
        painter.setPen(pen)
        angle = int(360 * self.value / 100)
        painter.drawArc(10, 10, 180, 180, 90 * 16, -angle * 16)
        painter.setPen(QColor("white"))
        painter.setFont(QFont("Segoe UI", 25, QFont.Weight.Bold))
        painter.drawText(self.rect(), Qt.AlignmentFlag.AlignCenter, f"%{int(self.value)}")


class ODin3DPC_GUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("ODin3DPC - Autonomous Vision System")
        self.setMinimumSize(1600, 950)
        self.setStyleSheet("background: #0d1117;")

        self.is_data_loaded = False
        self.pcd_data = None
        self.original_pcd_colors = None
        self.current_bboxes, self.current_labels = [], []
        self.box_items = []
        self.selected_idx, self.flash_state = -1, False
        self.is_heatmap_active = False
        self.is_segmentation_active = False

        self.flash_timer = QTimer()
        self.flash_timer.timeout.connect(self.toggle_flash)
        self.loading_timer = QTimer()
        self.loading_timer.timeout.connect(self.smooth_progress)
        self.target_value = 0

        self.init_ui()

    def init_ui(self):
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.layout_stack = QStackedLayout(self.central_widget)
        self.main_container = QWidget()
        main_layout = QHBoxLayout(self.main_container)

        left_p = QVBoxLayout()
        self.status = QLabel("SISTEM HAZIR - BEKLEMEDE")
        self.status.setStyleSheet("color: #58a6ff; font-weight: bold; font-family: 'Segoe UI'; font-size: 13pt;")
        left_p.addWidget(self.status)

        self.table = QTableWidget(0, 3)
        self.table.setHorizontalHeaderLabels(["ID", "Tip", "Mesafe"])
        self.table.setStyleSheet(
            "QTableWidget { background: #161b22; color: white; border: 1px solid #30363d; font-family: 'Segoe UI'; }")
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.itemClicked.connect(self.focus_on_object)
        left_p.addWidget(self.table)

        self.analysis_card = QGroupBox("Yapay Zeka Karar Analizi")
        self.analysis_card.setStyleSheet(
            "QGroupBox { color: #58a6ff; font-weight: bold; border: 1px solid #30363d; margin-top: 15px; padding-top: 20px; }")
        al = QVBoxLayout()
        al.setContentsMargins(15, 15, 15, 15)
        al.setSpacing(10)
        self.lbl_reason = QLabel("Detay icin bir nesne secin")
        self.lbl_reason.setWordWrap(True)
        self.lbl_reason.setStyleSheet("color: #c9d1d9; font-size: 10pt; line-height: 150%;")
        al.addWidget(self.lbl_reason)
        self.analysis_card.setLayout(al)
        left_p.addWidget(self.analysis_card)

        self.lbl_legend = QLabel("")
        self.lbl_legend.setStyleSheet("color: #8b949e; font-size: 8pt; font-family: 'Segoe UI';")
        left_p.addWidget(self.lbl_legend)

        btn_style = "QPushButton { background: #21262d; color: white; border: 1px solid #30363d; border-radius: 6px; padding: 12px; font-weight: bold; font-family: 'Segoe UI'; }"

        self.btn_segment = QPushButton("SEMANTIK SEGMENTASYON")
        self.btn_segment.setCheckable(True)
        self.btn_segment.setStyleSheet(btn_style.replace("#21262d", "#1d2d3d"))
        self.btn_segment.toggled.connect(self.toggle_segmentation)
        left_p.addWidget(self.btn_segment)

        self.btn_heatmap = QPushButton("ISI HARITASI MODU")
        self.btn_heatmap.setCheckable(True)
        self.btn_heatmap.setStyleSheet(btn_style)
        self.btn_heatmap.toggled.connect(self.toggle_heatmap)
        left_p.addWidget(self.btn_heatmap)

        self.btn_reset = QPushButton("KAMERA PERSPEKTIFINI SIFIRLA")
        self.btn_reset.setStyleSheet(btn_style)
        self.btn_reset.clicked.connect(self.reset_to_camera_view)
        left_p.addWidget(self.btn_reset)

        self.btn_load = QPushButton("VERI SETI YUKLE")
        self.btn_load.setStyleSheet(btn_style.replace("#21262d", "#238636"))
        self.btn_load.clicked.connect(self.start_loading)
        left_p.addWidget(self.btn_load)

        self.btn_export = QPushButton("ANALIZ RAPORUNU INDIR (CSV)")
        self.btn_export.setStyleSheet(btn_style.replace("#21262d", "#8957e5"))
        self.btn_export.clicked.connect(self.export_csv)
        left_p.addWidget(self.btn_export)

        main_layout.addLayout(left_p, 1)

        # Tüm butonlar yaratıldıktan SONRA başlangıç için pasif hale getiriyoruz
        self.btn_segment.setEnabled(False)
        self.btn_heatmap.setEnabled(False)
        self.btn_reset.setEnabled(False)
        self.btn_export.setEnabled(False)

        right_p = QVBoxLayout()

        self.btn_segment.setEnabled(False)
        self.btn_heatmap.setEnabled(False)
        self.btn_reset.setEnabled(False)

        right_p = QVBoxLayout()
        self.splitter = QSplitter(Qt.Orientation.Vertical)
        self.cam_lbl = QLabel("NEURAL-VISION [2D]")
        self.cam_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.cam_lbl.setStyleSheet("background: black; border: 1px solid #30363d;")

        self.lidar_view = gl.GLViewWidget()
        self.lidar_view.setBackgroundColor('#050505')

        self.splitter.addWidget(self.cam_lbl)
        self.splitter.addWidget(self.lidar_view)
        self.splitter.setSizes([450, 450])
        self.splitter.setStyleSheet("QSplitter::handle { background: #30363d; }")

        right_p.addWidget(self.splitter)
        main_layout.addLayout(right_p, 3)

        self.loading_overlay = QWidget()
        self.loading_overlay.setStyleSheet("background: rgba(13, 17, 23, 220);")
        loading_layout = QVBoxLayout(self.loading_overlay)
        self.loader = CircularProgress()
        loading_layout.addStretch()
        loading_layout.addWidget(self.loader, 0, Qt.AlignmentFlag.AlignCenter)
        loading_layout.addStretch()

        self.layout_stack.addWidget(self.main_container)
        self.layout_stack.addWidget(self.loading_overlay)

    def smooth_progress(self):
        if self.loader.value < self.target_value:
            self.loader.value += 1
            self.loader.update()
        elif self.loader.value >= 100:
            self.loading_timer.stop()
            self.layout_stack.setCurrentIndex(0)
            self.loading_overlay.hide()

    def start_loading(self):
        # YENI: Windows'un karisik siralamasini devre disi birakip ozel Qt penceresi aciyoruz
        dialog = QFileDialog(self, "LiDAR Dosyasi Sec", "data/velodyne/")
        dialog.setNameFilter("BIN Files (*.bin)")
        dialog.setOption(QFileDialog.Option.DontUseNativeDialog, True)  # Windows penceresini iptal et

        if dialog.exec():
            path = dialog.selectedFiles()[0]
        else:
            return

        self.current_filename = os.path.basename(path)

        self.btn_segment.setEnabled(False)
        self.btn_heatmap.setEnabled(False)
        self.btn_reset.setEnabled(False)
        self.btn_export.setEnabled(False)

        self.is_data_loaded = False
        self.flash_timer.stop()
        self.selected_idx = -1
        self.lbl_reason.setText("Detay icin bir nesne secin")

        self.loader.value = 0
        self.target_value = 85
        self.loading_overlay.show()
        self.layout_stack.setCurrentIndex(1)
        self.loading_timer.start(20)

        self.worker = DataLoaderThread(path)
        self.worker.finished_signal.connect(self.on_load_success)
        self.worker.error_signal.connect(self.on_load_error)
        self.worker.start()

    def on_load_success(self, points_np, safe_bboxes, meta, img_path):
        try:
            img_res = None
            if os.path.exists(img_path):
                img = cv2.imread(img_path)
                if img is not None:
                    img_res = detect_objects_2d(img)

            self.pcd_data = o3d.geometry.PointCloud()
            self.pcd_data.points = o3d.utility.Vector3dVector(points_np)
            self.original_pcd_colors = np.ones((points_np.shape[0], 3), dtype=np.float32) * 0.5

            combined = sorted(zip(safe_bboxes, meta), key=lambda x: np.linalg.norm(x[0]['center']))
            self.current_bboxes = []
            self.current_labels = [x[1] for x in combined]

            for b_dict in [x[0] for x in combined]:
                b = o3d.geometry.AxisAlignedBoundingBox(b_dict['min_bound'], b_dict['max_bound'])
                b.color = b_dict['color']
                self.current_bboxes.append(b)

            self.update_table()

            self.lidar_view.clear()
            self.box_items = []

            points = np.asarray(self.pcd_data.points, dtype=np.float32)
            rgba_colors = np.ones((points.shape[0], 4), dtype=np.float32)
            rgba_colors[:, :3] = self.original_pcd_colors
            self.scatter = gl.GLScatterPlotItem(pos=points, color=rgba_colors, size=1.5)
            self.lidar_view.addItem(self.scatter)

            for i, b in enumerate(self.current_bboxes):
                w, l, h = b.get_extent()
                center = b.get_center()
                c = b.color
                # HATANIN CÖZÜMÜ: Kutularin boyutu icin de QVector3D kullanildi
                box = gl.GLBoxItem(size=QVector3D(float(w), float(l), float(h)),
                                   color=(int(c[0]*255), int(c[1]*255), int(c[2]*255), 255))
                box.translate(float(center[0] - w/2), float(center[1] - l/2), float(center[2] - h/2))
                self.lidar_view.addItem(box)
                self.box_items.append(box)

            if img_res is not None:
                rgb = cv2.cvtColor(img_res, cv2.COLOR_BGR2RGB)
                qimg = QImage(rgb.data, rgb.shape[1], rgb.shape[0], rgb.shape[1]*3, QImage.Format.Format_RGB888)
                self.cam_lbl.setPixmap(QPixmap.fromImage(qimg).scaled(self.cam_lbl.size(), Qt.AspectRatioMode.KeepAspectRatio))
            else:
                self.cam_lbl.setText("ESLESTIRILECEK FOTOGRAF BULUNAMADI")

            self.reset_to_camera_view()
            self.is_data_loaded = True

            self.btn_segment.setEnabled(True)
            self.btn_heatmap.setEnabled(True)
            self.btn_reset.setEnabled(True)
            self.btn_export.setEnabled(True)

            self.target_value = 100
            self.status.setText("VERI BASARIYLA YUKLENDI")
            self.lidar_view.update()
        except Exception as e:
            self.on_load_error(str(e))

    def on_load_error(self, error_msg):
        self.loading_timer.stop()
        self.layout_stack.setCurrentIndex(0)
        self.status.setText("YUKLEME HATASI!")
        QMessageBox.critical(self, "Hata", f"Veri islenirken hata olustu:\n{error_msg}")

    def update_table(self):
        self.table.setRowCount(0)
        self.table.setColumnCount(3)  # Sadece 3 sutun kullanilacak
        for i, (b, meta) in enumerate(zip(self.current_bboxes, self.current_labels)):
            r = self.table.rowCount()
            self.table.insertRow(r)
            dist = np.linalg.norm(b.get_center())
            tip = meta.split("|")[0].strip()  # Hiz verisini yoksayip sadece Tipi aliyoruz

            self.table.setItem(r, 0, QTableWidgetItem(f"ID {i}"))
            self.table.setItem(r, 1, QTableWidgetItem(tip))
            self.table.setItem(r, 2, QTableWidgetItem(f"{dist:.2f} m"))

    def export_csv(self):
        if not self.is_data_loaded: return
        import pandas as pd
        import os
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        user_path = os.path.expanduser("~")
        possible_paths = [
            os.path.join(user_path, "OneDrive", "Masaüstü"),
            os.path.join(user_path, "OneDrive", "Desktop"),
            os.path.join(user_path, "Masaüstü"),
            os.path.join(user_path, "Desktop")
        ]

        desktop_path = possible_paths[-1]
        for p in possible_paths:
            if os.path.exists(p):
                desktop_path = p
                break

        clean_name = self.current_filename.replace('.bin', '')
        default_filename = f"{clean_name} veri seti nesne analizi tablosu.xlsx"
        default_path = os.path.join(desktop_path, default_filename)

        path, _ = QFileDialog.getSaveFileName(self, "Excel Raporu Kaydet", default_path, "Excel Files (*.xlsx)", options=QFileDialog.Option.DontUseNativeDialog)
        if not path: return

        try:
            data = []
            for i, (b, meta) in enumerate(zip(self.current_bboxes, self.current_labels)):
                tip = meta.split("|")[0].strip()  # Hiz verisini iceri almiyoruz
                w, l, h = b.get_extent()
                dist = np.linalg.norm(b.get_center())

                data.append({
                    "Nesne ID": f"ID {i}",
                    "Tespit Edilen Sinif": tip,
                    "Genislik (m)": round(w, 2),
                    "Uzunluk (m)": round(l, 2),
                    "Yukseklik (m)": round(h, 2),
                    "Kameraya Olan Mesafe (m)": round(dist, 2)  # İsim "Kameraya" olarak degistirildi
                })

            df = pd.DataFrame(data)

            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Analiz Raporu')
                worksheet = writer.sheets['Analiz Raporu']

                # Renk, Font ve Hizalama ayarlari
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                center_alignment = Alignment(horizontal="center", vertical="center")

                light_blue_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
                white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

                # Tabloları çerçevelemek için Siyah Kenarlık Stili (YENİ)
                thin_border = Border(left=Side(style='thin', color='000000'),
                                     right=Side(style='thin', color='000000'),
                                     top=Side(style='thin', color='000000'),
                                     bottom=Side(style='thin', color='000000'))

                # Ana Tablo Başlıkları
                for col_num, column_title in enumerate(df.columns.values):
                    col_letter = get_column_letter(col_num + 1)
                    max_len = max(df[column_title].astype(str).map(len).max(), len(column_title))
                    worksheet.column_dimensions[col_letter].width = max_len + 5

                    header_cell = worksheet.cell(row=1, column=col_num + 1)
                    header_cell.fill = header_fill
                    header_cell.font = header_font
                    header_cell.alignment = center_alignment
                    header_cell.border = thin_border  # Kenarlık eklendi

                # Ana Tablo Verileri
                for row_idx, row in enumerate(
                        worksheet.iter_rows(min_row=2, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)),
                        start=2):
                    current_fill = light_blue_fill if row_idx % 2 == 0 else white_fill
                    for col_idx, cell in enumerate(row):
                        # Eger ilk sutunsa (Nesne ID), baslik rengi (Koyu Mavi) yap
                        if col_idx == 0:
                            cell.fill = header_fill
                            cell.font = header_font
                        else:
                            cell.fill = current_fill

                        cell.alignment = center_alignment
                        cell.border = thin_border  # Kenarlık eklendi

                # --- 3. ISTATISTIK TABLOSU ---
                stats_start_row = len(df) + 7

                header1 = worksheet.cell(row=stats_start_row, column=1, value="ISTATISTIKSEL OZET")
                header1.fill = header_fill
                header1.font = header_font
                header1.alignment = center_alignment
                header1.border = thin_border

                header2 = worksheet.cell(row=stats_start_row, column=2, value="DEGER")
                header2.fill = header_fill
                header2.font = header_font
                header2.alignment = center_alignment
                header2.border = thin_border

                total_obj = len(df)
                avg_dist = df["Kameraya Olan Mesafe (m)"].mean() if total_obj > 0 else 0
                class_counts = df["Tespit Edilen Sinif"].value_counts().to_dict()

                stats_data = [
                    ("Toplam Tespit Edilen Nesne", total_obj),
                    ("Ortalama Nesne Mesafesi (m)", round(avg_dist, 2))
                ]
                for cls, count in class_counts.items():
                    stats_data.append((f"Toplam '{cls}' Sayisi", count))

                # İstatistik tablosunun A sütununun genisligini yazi kesilmeyecek sekilde sabitliyoruz
                worksheet.column_dimensions['A'].width = max(worksheet.column_dimensions['A'].width, 35)

                # İstatistik verilerini yazma
                for i, (stat_name, stat_val) in enumerate(stats_data):
                    current_row = stats_start_row + 1 + i
                    c1 = worksheet.cell(row=current_row, column=1, value=stat_name)
                    c2 = worksheet.cell(row=current_row, column=2, value=stat_val)

                    current_fill = light_blue_fill if i % 2 == 0 else white_fill

                    # Istatistik satir basliklari ID gibi koyu mavi ve cerceveli olacak
                    c1.fill = header_fill
                    c1.font = header_font
                    c1.alignment = Alignment(horizontal="left", vertical="center")
                    c1.border = thin_border

                    # Degerler acik mavi/beyaz oruntulu ve cerceveli olacak
                    c2.fill = current_fill
                    c2.alignment = center_alignment
                    c2.border = thin_border

            QMessageBox.information(self, "Basarili",
                                    f"Rapor profesyonel bir Excel tablosu olarak kaydedildi!\n\nKonum: {path}")

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Rapor kaydedilemedi: {str(e)}")

    def toggle_flash(self):
        if not self.is_data_loaded or self.selected_idx == -1: return
        target_box = self.box_items[self.selected_idx]
        tip = self.current_labels[self.selected_idx].split("|")[0].strip()
        c = [1, 0, 0] if "Arac" in tip or "Kamyon" in tip else ([0, 1, 0] if "Yaya" in tip else [0.5, 0.5, 0.5])

        self.lidar_view.removeItem(target_box)
        w, l, h = self.current_bboxes[self.selected_idx].get_extent()
        center = self.current_bboxes[self.selected_idx].get_center()

        new_color = (255, 255, 255, 255) if self.flash_state else (int(c[0]*255), int(c[1]*255), int(c[2]*255), 255)
        # HATANIN CÖZÜMÜ: Kutularin boyutu icin de QVector3D kullanildi
        new_box = gl.GLBoxItem(size=QVector3D(float(w), float(l), float(h)), color=new_color)
        new_box.translate(float(center[0]-w/2), float(center[1]-l/2), float(center[2]-h/2))
        self.lidar_view.addItem(new_box)
        self.box_items[self.selected_idx] = new_box
        self.flash_state = not self.flash_state
        self.lidar_view.update()

    def reset_colors(self):
        if self.pcd_data and self.original_pcd_colors is not None:
            points = np.asarray(self.pcd_data.points, dtype=np.float32)
            rgba_colors = np.ones((points.shape[0], 4), dtype=np.float32)
            rgba_colors[:, :3] = self.original_pcd_colors
            self.scatter.setData(color=rgba_colors)

        for i, box in enumerate(self.box_items):
            self.lidar_view.removeItem(box)
            w, l, h = self.current_bboxes[i].get_extent()
            center = self.current_bboxes[i].get_center()
            tip = self.current_labels[i].split("|")[0].strip()
            c = [1, 0, 0] if "Arac" in tip or "Kamyon" in tip else ([0, 1, 0] if "Yaya" in tip else [0.5, 0.5, 0.5])
            # HATANIN CÖZÜMÜ: Kutularin boyutu icin de QVector3D kullanildi
            new_box = gl.GLBoxItem(size=QVector3D(float(w), float(l), float(h)), color=(int(c[0]*255), int(c[1]*255), int(c[2]*255), 255))
            new_box.translate(float(center[0]-w/2), float(center[1]-l/2), float(center[2]-h/2))
            self.lidar_view.addItem(new_box)
            self.box_items[i] = new_box
        self.lbl_legend.setText("")
        self.lidar_view.update()

    def toggle_segmentation(self, checked):
        if not self.is_data_loaded: return
        self.is_segmentation_active = checked
        points = np.asarray(self.pcd_data.points, dtype=np.float32)
        rgba = np.ones((points.shape[0], 4), dtype=np.float32)

        if checked:
            self.btn_heatmap.setChecked(False)
            colors = np.ones((points.shape[0], 3), dtype=np.float32) * 0.2
            for i, box in enumerate(self.current_bboxes):
                indices = box.get_point_indices_within_bounding_box(self.pcd_data.points)
                tip = self.current_labels[i].split("|")[0].strip()
                color = [1, 0, 0] if "Arac" in tip else ([0, 1, 0] if "Yaya" in tip else [0.2, 0.8, 0.2] if "Agac" in tip else [0.5, 0.5, 0.8])
                colors[indices] = color
            rgba[:, :3] = colors
            self.lbl_legend.setText("Segmentasyon: Kirmizi (Arac), Yesil (Yaya/Agac), Gri (Zemin)")
        else:
            rgba[:, :3] = self.original_pcd_colors
            if not self.is_heatmap_active: self.lbl_legend.setText("")
        self.scatter.setData(color=rgba)
        self.lidar_view.update()

    def toggle_heatmap(self, checked):
        if not self.is_data_loaded: return
        self.is_heatmap_active = checked
        points = np.asarray(self.pcd_data.points, dtype=np.float32)
        rgba = np.ones((points.shape[0], 4), dtype=np.float32)

        if checked:
            self.btn_segment.setChecked(False)
            z_vals = points[:, 2]
            # Z (Yukseklik) degerlerini 0 ile 1 arasina normalize et
            z_norm = (z_vals - z_vals.min()) / (z_vals.max() - z_vals.min() + 1e-8)
            colors = np.zeros((points.shape[0], 3), dtype=np.float32)

            # --- COGRAFI RENK SKALASI MATEMATIGI ---
            c1 = z_norm < 0.25
            c2 = (z_norm >= 0.25) & (z_norm < 0.50)
            c3 = (z_norm >= 0.50) & (z_norm < 0.75)
            c4 = z_norm >= 0.75

            # 1. Seviye: Mavi'den Yesil'e (Zemin ve cukurlar)
            colors[c1, 1] = z_norm[c1] / 0.25  # Yesil artar
            colors[c1, 2] = 1 - (z_norm[c1] / 0.25)  # Mavi azalir

            # 2. Seviye: Yesil'den Sari'ya (Arac lastikleri, kaldirimlar)
            colors[c2, 0] = (z_norm[c2] - 0.25) / 0.25  # Kirmizi artar
            colors[c2, 1] = 1.0  # Yesil full

            # 3. Seviye: Sari'dan Turuncu'ya (Arac tavanlari, tabelalar)
            colors[c3, 0] = 1.0  # Kirmizi full
            colors[c3, 1] = 1.0 - 0.5 * ((z_norm[c3] - 0.50) / 0.25)  # Yesil yariya duser

            # 4. Seviye: Turuncu'dan Kirmizi'ya (Agac tepeleri, binalar)
            colors[c4, 0] = 1.0  # Kirmizi full
            colors[c4, 1] = 0.5 - 0.5 * ((z_norm[c4] - 0.75) / 0.25)  # Yesil tamamen biter

            rgba[:, :3] = colors
            self.lbl_legend.setText("Cografi Isı Haritasi: Mavi (Zemin) -> Yesil -> Sari -> Kirmizi (Tepe)")
        else:
            rgba[:, :3] = self.original_pcd_colors
            if not self.is_segmentation_active: self.lbl_legend.setText("")

        self.scatter.setData(color=rgba)
        self.lidar_view.update()

    def reset_to_camera_view(self):
        if not self.is_data_loaded: return

        # Bakış merkezini tam olarak aracın 15 metre önüne sabitliyoruz
        self.lidar_view.opts['center'] = QVector3D(0.0, 0.0, 15.0)

        # distance: Kameranın merkeze olan uzaklığı
        # elevation: 0 (Tam karşıya bakış, ufuk çizgisi)
        # azimuth: -90 (Tam ileri bakış yönü)
        self.lidar_view.setCameraPosition(distance=18.0, elevation=0.0, azimuth=-90.0)

        # FOV değerini 60-70 bandına çekmek derinlik algısını daha gerçekçi yapar
        self.lidar_view.opts['fov'] = 70
        self.lidar_view.update()

    def reset_single_box_color(self, idx):
        if idx == -1 or idx >= len(self.box_items): return
        self.lidar_view.removeItem(self.box_items[idx])
        w, l, h = self.current_bboxes[idx].get_extent()
        center = self.current_bboxes[idx].get_center()
        tip = self.current_labels[idx].split("|")[0].strip()
        c = [1, 0, 0] if "Arac" in tip or "Kamyon" in tip else ([0, 1, 0] if "Yaya" in tip else [0.5, 0.5, 0.5])
        # HATANIN CÖZÜMÜ: Kutularin boyutu icin de QVector3D kullanildi
        new_box = gl.GLBoxItem(size=QVector3D(float(w), float(l), float(h)), color=(int(c[0]*255), int(c[1]*255), int(c[2]*255), 255))
        new_box.translate(float(center[0]-w/2), float(center[1]-l/2), float(center[2]-h/2))
        self.lidar_view.addItem(new_box)
        self.box_items[idx] = new_box
        self.lidar_view.update()

    def focus_on_object(self, item):
        if self.selected_idx != -1:
            self.reset_single_box_color(self.selected_idx)
            self.flash_timer.stop()

        self.selected_idx = item.row()
        self.flash_state = False

        target = self.current_bboxes[self.selected_idx]
        tip = self.current_labels[self.selected_idx].split("|")[0].strip()
        w, l, h = target.get_extent()
        center = target.get_center()

        # HATANIN CÖZÜMÜ: Kamera merkez noktalari QVector3D olarak ayarlandi
        self.lidar_view.opts['center'] = QVector3D(float(center[0]), float(center[1]), float(center[2]))
        self.lidar_view.opts['distance'] = 15.0
        self.lidar_view.update()

        analysis = f"<b>Tespit:</b> {tip}<br><br>Boyutlar: {w:.1f}m x {l:.1f}m x {h:.1f}m<br>"
        if "Arac" in tip or "Kamyon" in tip:
            analysis += "AI Karari: Karayolu tasiti geometrisiyle tam uyusuyor."
        elif "Bina" in tip:
            analysis += "AI Karari: Buyuk hacim ve yukseklik nedeniyle sabit yapi olarak tanindi."
        elif "Agac" in tip:
            analysis += "AI Karari: Ince dikey form ve duzensiz ust yapi agaca isaret ediyor."
        else:
            analysis += "AI Karari: Standart disi boyut; engel olarak isaretlendi."

        self.lbl_reason.setText(analysis)
        self.flash_timer.start(400)